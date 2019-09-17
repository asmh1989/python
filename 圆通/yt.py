#!/usr/bin/python3

from openpyxl import Workbook, load_workbook
import string
import json

ysb = load_workbook(filename='./圆通原始表.xlsx')
print(ysb.sheetnames)
ws2 = ysb['Sheet2']
ws1 = ysb['Sheet1']

d = {}
i = 3
print('运费计算表: --------------------------------')

while True:
    loc = str(i)
    # 地区
    area = ws2['B' + loc].value

    if area == None:
        break
    i += 1

    # 5kg 以上 首重
    above5_kg = ws2['D' + loc].value
    # 5kg 以上续重
    above52_kg = ws2['E' + loc].value

    below5_1 = ws2['F' + loc].value
    below5_2 = ws2['G' + loc].value
    below5_3 = ws2['H' + loc].value
    below5_4 = ws2['I' + loc].value

    d[area] = [below5_1, below5_2, below5_3, below5_4, above5_kg, above52_kg]
    print(area + ' : ' + str(d[area]))

print('运费计算表: --------------------------------')


def cal(area, kg):
    kgs = d[area]

    z = int(kg)
    y = kg - z

    if kg < 1.0001:
        m = kgs[4]
    else:
        if y < 0.0001:
            m = kgs[4] + (z - 1.0) * kgs[5]
        else:
            m = kgs[4] + (z - 1.0 + (1 if y > 0.5 else 0.5)) * kgs[5]

    if kg < 1.0001:
        if kgs[0] != 0:
            m = kgs[0]
    elif kg < 3.0001:
        if kgs[1] != 0:
            m = kgs[1]
    elif kg < 4.0001:
        if kgs[1] != 0:
            m = kgs[2]
    elif kg < 5.0001:
        if kgs[3] != 0:
            m = kgs[3]

    return m


print('开始计算sheet1中的快递费...')

i = 2
while True:
    loc = str(i)
    area = ws1['E' + loc].value
    if area == None:
        break
    i += 1

    kg = ws1['C' + loc].value
    info = ws1['B' + loc].value

    money = cal(area, kg)

    ws1['G' + loc] = money

    ws1['H' + loc] = '=F' + loc + '-G' + loc

    # print(info + ' 运费 ===> ' + str(money))

last = str(i - 1)
ws1['G' + str(i)] = "=SUM(G2:G" + last + ")"
ysb.save('圆通测算表.xlsx')

print('完成!!')
