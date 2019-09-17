#!/usr/bin/python3

from openpyxl import Workbook, load_workbook
import string
import json

ysb = load_workbook(filename='./邮政原始表.xlsx')
print(ysb.sheetnames)
ws2 = ysb['Sheet1']
ws1 = ysb['邮件查询']

d = {}
i = 3
print('运费计算表: --------------------------------')

while True:
    loc = str(i)
    # 地区
    area = ws2['B' + loc].value

    if area == None:
        break
    i += 1

    # 5kg 以上 首重
    above5_kg = ws2['J' + loc].value
    # 5kg 以上续重
    above52_kg = ws2['K' + loc].value

    below5_1 = ws2['C' + loc].value

    below5_2 = ws2['D' + loc].value

    below5_3 = ws2['F' + loc].value
    below52_3 = ws2['G' + loc].value
    below5_4 = ws2['H' + loc].value
    below52_4 = ws2['I' + loc].value

    d[area] = [
        below5_1, below5_2, below5_3, below52_3, below5_4, below52_4,
        above5_kg, above52_kg
    ]
    print(area + ' : ' + str(d[area]))

print('运费计算表: --------------------------------')


def cal(area, k):
    kgs = d[area]

    kg = k / 1000.0
    z = int(kg)
    y = kg - z

    if y > 0.001:
        z += 1

    if kg < 0.5001:
        m = kgs[0]
    elif kg < 1.0001:
        m = kgs[1]
    elif kg < 3.0001:
        m = kgs[2] + (z - 1.0) * kgs[3]
    elif kg < 5.0001:
        m = kgs[4] + (z - 1.0) * kgs[5]
    else:
        m = kgs[6] + (z - 1.0) * kgs[7]

    return m


print('开始计算sheet1中的快递费...')

i = 2
while True:
    loc = str(i)
    area = ws1['C' + loc].value
    if area == None:
        break
    if str(area) < '1':
        break
    i += 1

    kg = ws1['F' + loc].value
    info = ws1['B' + loc].value

    money = cal(area, kg)

    ws1['H' + loc] = money

    ws1['I' + loc] = '=G' + loc + '-H' + loc

    print(info + ' 运费 ===> ' + str(money))

last = str(i - 1)
ws1['H' + str(i)] = "=SUM(H2:H" + last + ")"
ysb.save('邮政测算表.xlsx')

print('完成!!')
